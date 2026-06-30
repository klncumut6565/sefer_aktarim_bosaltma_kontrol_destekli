#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PDF Sefer Raporu → Excel Taşıma Kontrol Listesi + Boşaltma Kontrol Dökümanı
Çekirdek iş mantığı (UI bağımsız). Hem masaüstü (Tkinter) hem web (Streamlit)
arayüzleri tarafından ortak kullanılır.

Gereksinimler:
    pip install openpyxl pdfplumber python-docx
"""

from __future__ import annotations

import io
import logging
import os
import re
import sys
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection
from openpyxl.drawing.image import Image as XLImage

try:
    from docx_doldur import kontrol_dokumani_olustur
    _DOCX_DESTEGI = True
except Exception:
    _DOCX_DESTEGI = False


def _resource_path(relative_path: str) -> Path:
    """Hem `python sefer_aktarim_zebra.py` ile hem de PyInstaller exe içinde
    çalışırken aynı dosyayı (örn. logo.ico) doğru konumdan bulur."""
    base_path = getattr(sys, "_MEIPASS", None)
    if base_path is None:
        base_path = Path(__file__).resolve().parent
    return Path(base_path) / relative_path

# ---------------------------------------------------------------------------
# Loglama (hem konsola hem arayüze)
# ---------------------------------------------------------------------------
log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-8s %(message)s", datefmt="%H:%M:%S")

# ---------------------------------------------------------------------------
# Sabitler
# ---------------------------------------------------------------------------
SCRIPT_DIR     = Path(__file__).parent
DATA_START_ROW = 6
HEADER_ROW     = 5
MAX_COL        = 22

ZEBRA_COLORS = [
    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"),
]

COL = {
    "sira_no": 1, "tasima_tarihi": 2, "tasiyici_unvan": 3,
    "gonderen": 4, "alici": 5, "vergi_no": 6, "plaka": 7,
    "tasima_turu": 8, "sefer_no": 9, "un_no": 10, "urun_adi": 11,
    "miktar": 12, "birim": 13, "src_sofor": 20, "muayene_tarihi": 16,
    "muafiyet": 19,
}


# ---------------------------------------------------------------------------
# Veri modeli
# ---------------------------------------------------------------------------
@dataclass
class Yuk:
    sefer_no: str; plaka: str; tarih: Optional[datetime]; src_sofor: str
    tasiyici_unvan: str; gonderen_vn: str; gonderen: str
    alici_vn: str; alici: str; un_no: str; urun_adi: str; miktar: int
    birim: str = "Lt"; tasima_turu: str = "ADR-TANK"
    plaka_sabit: str = ""           # PDF'den okunan orijinal plaka
    muayene_tarihi: Optional[datetime] = None
    muafiyet: str = ""
    kontrol_dokumani_yolu: str = ""   # Sefer No hücresine eklenecek köprünün hedef yolu (göreceli)

    @property
    def sefer_no_int(self):
        try: return int(self.sefer_no)
        except ValueError: return 0

    def sort_key(self):
        return (self.tarih or datetime.min, self.sefer_no_int, self.urun_adi)

    def to_excel_map(self, sira_no: int) -> dict:
        sefer_val = int(self.sefer_no) if self.sefer_no.isdigit() else self.sefer_no
        un_val    = int(self.un_no)    if self.un_no.isdigit()    else self.un_no
        # UN No 1203 ise ürün adı BENZIN olarak sabitlenir
        urun = "BENZIN" if self.un_no == "1203" else self.urun_adi
        # Plaka: sabit plaka varsa onu kullan (artırılmamış orijinal)
        plaka_val = self.plaka_sabit if self.plaka_sabit else self.plaka
        # Muafiyet metni
        muafiyet_text = self.muafiyet if self.muafiyet else self._default_muafiyet()
        return {
            COL["sira_no"]: sira_no, COL["tasima_tarihi"]: self.tarih.date() if self.tarih else None,
            COL["tasiyici_unvan"]: self.tasiyici_unvan, COL["gonderen"]: self.gonderen,
            COL["alici"]: self.alici, COL["vergi_no"]: self.gonderen_vn,
            COL["plaka"]: plaka_val, COL["tasima_turu"]: self.tasima_turu,
            COL["sefer_no"]: sefer_val, COL["un_no"]: un_val,
            COL["urun_adi"]: urun, COL["miktar"]: self.miktar,
            COL["birim"]: self.birim, COL["src_sofor"]: self.src_sofor,
            COL["muayene_tarihi"]: self.muayene_tarihi.date() if self.muayene_tarihi else None,
            COL["muafiyet"]: muafiyet_text,
        }

    def _default_muafiyet(self) -> str:
        """UN No ve miktara göre muafiyet metni.
        ADR 1.1.3.6 taşıma kategorisi limit değerleri:
          UN 1203 (Taşıma Kategorisi 2): sınır 333 (Lt/Kg)
          UN 1202 (Taşıma Kategorisi 3): sınır 1000 (Lt/Kg)
        Miktar sınır değere eşit veya altındaysa EVET (muafiyet kapsamında),
        sınırı aşıyorsa HAYIR (muafiyet kapsamı dışında) yazılır.
        """
        if self.un_no == "1203":
            kapsam = "EVET" if self.miktar <= 333 else "HAYIR"
            return f"{kapsam}\n-TAŞIMA KATEGORİSİ-2\nSINIR 333"
        elif self.un_no == "1202":
            kapsam = "EVET" if self.miktar <= 1000 else "HAYIR"
            return f"{kapsam}\n-TAŞIMA KATEGORİSİ-3\nSINIR 1000"
        return ""


# ---------------------------------------------------------------------------
# PDF çıkarma (pdfplumber ile - pdftotext gerektirmez)
# ---------------------------------------------------------------------------
def _re(pattern, text, group=1, flags=0):
    m = re.search(pattern, text, flags)
    return m.group(group).strip() if m else ""

def extract_pdf_data(pdf_path: Path, tasima_turu_secimi: str = "ADR-AMBALAJLI") -> tuple[list[Yuk], str]:
    """PDF'den veri çıkarır. (yükler_listesi, plaka_sabit) döner.

    tasima_turu_secimi: Arayüzde seçilen taşıma türü. UN 1202/1203 için
    seçimden bağımsız olarak her zaman ADR-TANK kullanılır.
    """
    if not pdf_path.is_file():
        raise FileNotFoundError(f"PDF bulunamadı: {pdf_path}")

    try:
        import pdfplumber
    except ImportError:
        raise ImportError("pdfplumber kurulu değil. Kurulum: pip install pdfplumber")

    text = ""
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"

    if not text.strip():
        return [], ""

    sefer_no = _re(r"SEFER NO\s+(\d+)", text)
    plaka    = _re(r"PLAKA\s+(.+?)\s+BELGE", text)
    # Plaka sabit: PDF'den okunan orijinal plaka (örn: "39SG206 / 39ACV713")
    plaka_sabit = plaka
    src      = _re(r"PERSONEL\s+(\d+)", text)
    f1       = _re(r"TAŞIYAN FİRMA\s+(.+?)\s+UNVAN", text, flags=re.DOTALL)
    f2       = _re(r"UNVAN\s+(.+)", text)
    unvan    = f"{f1} {f2}".strip()
    m_t = re.search(r"BAŞLANGIÇ\s+(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2})", text)
    tarih = datetime.strptime(m_t.group(1), "%d/%m/%Y %H:%M") if m_t else None
    m_g = re.search(r"Gönderen\s+\((\d+)\)\s+(.+)", text)
    gvn, g = (m_g.group(1), m_g.group(2).strip()) if m_g else ("", "")
    m_a = re.search(r"Alıcı\s+\((\d+)\)\s+(.+)", text)
    avn, a = (m_a.group(1), m_a.group(2).strip()) if m_a else ("", "")
    # FIX: KG / LT / Litre / TON + UN No boşluklu/boşluksuz + ondalıklı sayı desteği
    matches = re.findall(
        r"Yük Detayı\s+(.+?)(?:\(UN No:\s*(\d+)\))?\s*-\s*([\d.,]+)\s*"
        r"(KG|Kg|kg|LT|Lt|lt|L\b|Litre|LİTRE|litre|TON|Ton|ton)",
        text, re.DOTALL
    )
    if not matches:
        return [], plaka_sabit
    yukler = []
    for ur, un, mik_str, birim_ham in matches:
        urun  = " ".join(re.sub(r"\s*\(UN No:\s*\d+\)\s*", "", ur).split())
        un_no = un or ""
        # Birimi standartlaştır
        birim_upper = birim_ham.upper()
        if birim_upper in ("LT", "LİTRE", "LITRE", "L"):
            birim = "Lt"
        elif birim_upper == "TON":
            birim = "Ton"
        else:
            birim = "KG"
        # Miktarı normalize et: "1.400" → 1400, "1400,5" → 1400
        try:
            mik = int(float(mik_str.replace(".", "").replace(",", ".")))
        except ValueError:
            mik = 0
        # Muafiyet metni — ADR 1.1.3.6 taşıma kategorisi limit değerleri.
        # Miktar sınır değere eşit veya altındaysa EVET (muafiyet kapsamında),
        # sınırı aşıyorsa HAYIR (muafiyet kapsamı dışında).
        if un_no == "1203":
            kapsam = "EVET" if mik <= 333 else "HAYIR"
            muafiyet = f"{kapsam}\n-TAŞIMA KATEGORİSİ-2\nSINIR 333"
        elif un_no == "1202":
            kapsam = "EVET" if mik <= 1000 else "HAYIR"
            muafiyet = f"{kapsam}\n-TAŞIMA KATEGORİSİ-3\nSINIR 1000"
        else:
            muafiyet = ""
        # Taşıma türü: UN 1202/1203 her zaman ADR-TANK, diğerleri arayüzden seçilen tür
        tasima_turu = "ADR-TANK" if un_no in ("1202", "1203") else tasima_turu_secimi
        yukler.append(Yuk(sefer_no=sefer_no, plaka=plaka, tarih=tarih, src_sofor=src,
                          tasiyici_unvan=unvan, gonderen_vn=gvn, gonderen=g,
                          alici_vn=avn, alici=a, un_no=un_no, urun_adi=urun, miktar=mik,
                          birim=birim, plaka_sabit=plaka_sabit, muafiyet=muafiyet,
                          tasima_turu=tasima_turu))
    return yukler, plaka_sabit


# ---------------------------------------------------------------------------
# Excel işlemleri
# ---------------------------------------------------------------------------
def _read_existing(ws):
    yukler, sefer_nolar = [], set()
    for row in range(DATA_START_ROW, ws.max_row + 1):
        raw = ws.cell(row=row, column=COL["sefer_no"]).value
        if raw is None: continue
        sn = str(raw)
        sefer_nolar.add(sn)
        tv = ws.cell(row=row, column=COL["tasima_tarihi"]).value
        def cv(c): return ws.cell(row=row, column=c).value
        try: mik = int(cv(COL["miktar"]) or 0)
        except: mik = 0
        mevcut_plaka = str(cv(COL["plaka"]) or "")
        muayene_t = cv(COL["muayene_tarihi"])
        muafiyet_val = str(cv(COL["muafiyet"]) or "")
        # Sefer No hücresinde mevcut bir köprü varsa koru (yeniden yazma sırasında kaybolmasın)
        sefer_no_cell = ws.cell(row=row, column=COL["sefer_no"])
        mevcut_kontrol_yolu = sefer_no_cell.hyperlink.target if sefer_no_cell.hyperlink else ""
        yukler.append(Yuk(
            sefer_no=sn, plaka=mevcut_plaka,
            tarih=tv if isinstance(tv, datetime) else None,
            src_sofor=str(cv(COL["src_sofor"]) or ""),
            tasiyici_unvan=str(cv(COL["tasiyici_unvan"]) or ""),
            gonderen_vn=str(cv(COL["vergi_no"]) or ""),
            gonderen=str(cv(COL["gonderen"]) or ""), alici_vn="",
            alici=str(cv(COL["alici"]) or ""),
            un_no=str(cv(COL["un_no"]) or ""),
            urun_adi=str(cv(COL["urun_adi"]) or ""),
            miktar=mik, birim=str(cv(COL["birim"]) or "Lt"),
            tasima_turu=str(cv(COL["tasima_turu"]) or "ADR-TANK"),
            plaka_sabit=mevcut_plaka,
            muayene_tarihi=muayene_t if isinstance(muayene_t, datetime) else None,
            kontrol_dokumani_yolu=mevcut_kontrol_yolu,
            muafiyet=muafiyet_val,
        ))
    return yukler, sefer_nolar

def _template_style(ws):
    tr = HEADER_ROW
    for row in range(ws.max_row, DATA_START_ROW - 1, -1):
        if ws.cell(row=row, column=COL["sefer_no"]).value is not None:
            tr = row; break
    s = {}
    for c in range(1, MAX_COL + 1):
        cell = ws.cell(row=tr, column=c)
        s[c] = {"font": copy(cell.font), "border": copy(cell.border),
                 "number_format": cell.number_format,
                 "protection": copy(cell.protection), "alignment": copy(cell.alignment)}
    return s

def _apply_style(ws, row, style):
    for c, a in style.items():
        cell = ws.cell(row=row, column=c)
        cell.font = copy(a["font"]); cell.border = copy(a["border"])
        cell.number_format = a["number_format"]
        cell.protection = copy(a["protection"]); cell.alignment = copy(a["alignment"])

def _zebra_map(yukler):
    cm, idx = {}, 0
    for y in yukler:
        if y.sefer_no not in cm:
            cm[y.sefer_no] = ZEBRA_COLORS[idx % len(ZEBRA_COLORS)]; idx += 1
    return cm

def _excel_logo_ekle(ws, logo_bytes: bytes) -> None:
    """Excel'in sol üst köşesindeki (A1:C4 birleşik) logo alanına, verilen
    logo görselini ekler/değiştirir. Önceki logo varsa kaldırılır."""
    # Önceki logo(lar) varsa kaldır (yeniden aktarımlarda üst üste binmesin)
    if hasattr(ws, "_images"):
        ws._images = [img for img in ws._images if getattr(img, "anchor", None) != "A1"]
    img = XLImage(io.BytesIO(logo_bytes))
    img.width = 140
    img.height = 70
    ws.add_image(img, "A1")


def _rewrite(ws, yukler, style):
    for row in range(ws.max_row, DATA_START_ROW - 1, -1):
        ws.delete_rows(row)
    cm = _zebra_map(yukler)
    for i, y in enumerate(yukler, 1):
        row = DATA_START_ROW + i - 1
        _apply_style(ws, row, style)
        for c, v in y.to_excel_map(i).items():
            cell = ws.cell(row=row, column=c)
            cell.value = v
            # Muayene tarihi: tarih değeri olarak sakla, GG.AA.YYYY göster
            if c == COL["muayene_tarihi"] and v is not None:
                cell.number_format = "DD.MM.YYYY"
            if c == COL["tasima_tarihi"] and v is not None:
                cell.number_format = "DD.MM.YYYY"   
            # Çok satırlı metinler için wrap_text
            if isinstance(v, str) and "\n" in v:
                cell.alignment = Alignment(wrapText=True)
        # Sefer No hücresine kontrol dökümanı köprüsü (varsa)
        if y.kontrol_dokumani_yolu:
            sefer_no_cell = ws.cell(row=row, column=COL["sefer_no"])
            sefer_no_cell.hyperlink = y.kontrol_dokumani_yolu
            sefer_no_cell.font = copy(sefer_no_cell.font)
            sefer_no_cell.font = Font(
                name=sefer_no_cell.font.name, size=sefer_no_cell.font.size,
                bold=sefer_no_cell.font.bold, italic=sefer_no_cell.font.italic,
                color="0563C1", underline="single",
            )
        fill = cm[y.sefer_no]
        for c in range(1, MAX_COL - 1):  # U ve V sütunları (21-22) renksiz bırakılır
            ws.cell(row=row, column=c).fill = fill
        ws.row_dimensions[row].height = 54.75

def docx_to_pdf(docx_path: Path, cikti_klasor: Path = None) -> Optional[Path]:
    """LibreOffice (soffice) kullanarak .docx dosyasını .pdf'e çevirir.
    LibreOffice kurulu değilse veya çeviri başarısız olursa None döner."""
    import shutil
    import subprocess

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return None

    hedef = cikti_klasor or docx_path.parent
    hedef.mkdir(parents=True, exist_ok=True)
    try:
        subprocess.run(
            [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(hedef), str(docx_path)],
            check=True, capture_output=True, timeout=60,
        )
    except Exception:
        return None
    pdf_yolu = hedef / (docx_path.stem + ".pdf")
    return pdf_yolu if pdf_yolu.is_file() else None


def process_pdfs(excel_path: Path, pdf_paths: list[Path], output_path: Path, log_cb=None, 
                                   muayene_tarihleri: dict[str, str] = None,
                                   tasima_turu: str = "ADR-AMBALAJLI",
                                   ek_tarihler: dict[str, dict] = None,
                                   docx_uret: bool = False,
                                   docx_cikti_klasor: Path = None,
                                   bosaltan_adi: str = "",
                                   sofor_adi: str = "",
                                   docx_pdf_donustur: bool = False,
                                   logo_bytes: Optional[bytes] = None) -> dict:
    """
    ek_tarihler: plaka -> {
        "yangin_tup": "GG.AA.YYYY", "tmfb": "...", "adr_uygunluk": "...",
        "periyodik_muayene": "..."   (ara_muayene zaten muayene_tarihleri içinde geliyor)
    }
    docx_uret: True ise her sefer için Boşaltma Kontrol Dökümanı (.docx) de üretilir.
    docx_cikti_klasor: docx dosyalarının kaydedileceği klasör (None ise excel ile aynı klasör).
    docx_pdf_donustur: True ise üretilen .docx ayrıca .pdf'e çevrilir (LibreOffice gerektirir,
        web/Streamlit ortamı için). Sonuçta hem docx hem pdf yolu döner.
    """
    def emit(msg):
        log.info(msg)
        if log_cb: log_cb(msg)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    style = _template_style(ws)
    mevcut, sefer_nolar = _read_existing(ws)

    yeni, atlanan, eklenen = [], 0, 0
    docx_uretilen = 0
    uretilen_dosyalar = []  # [{"sefer_no": ..., "docx": Path, "pdf": Path|None}, ...]
    for pdf in pdf_paths:
        emit(f"📄 İşleniyor: {pdf.name}")
        yukler, plaka_sabit = extract_pdf_data(pdf, tasima_turu_secimi=tasima_turu)
        if not yukler:
            emit(f"  ⚠ Yük bulunamadı, atlandı.")
            continue
        sn = yukler[0].sefer_no
        if sn in sefer_nolar:
            emit(f"  ⚠ Sefer {sn} zaten mevcut, atlandı.")
            atlanan += 1; continue

        # Muayene tarihi ekle (varsa)
        muayene_tarihi = None
        if muayene_tarihleri and plaka_sabit in muayene_tarihleri:
            mt_str = muayene_tarihleri[plaka_sabit]
            try:
                muayene_tarihi = datetime.strptime(mt_str, "%d.%m.%Y")
            except ValueError:
                muayene_tarihi = None
            for y in yukler:
                y.muayene_tarihi = muayene_tarihi

        yeni.extend(yukler); sefer_nolar.add(sn); eklenen += 1
        emit(f"  ✓ Sefer {sn} — {len(yukler)} satır eklenecek.")

        # ---- Boşaltma Kontrol Dökümanı (.docx / .pdf) üretimi ----
        if docx_uret and _DOCX_DESTEGI:
            try:
                ek = (ek_tarihler or {}).get(plaka_sabit, {})
                un_listesi = ", ".join(sorted({f"UN {y.un_no}" for y in yukler if y.un_no})) or "-"
                hedef_klasor = docx_cikti_klasor or output_path.parent
                hedef_klasor.mkdir(parents=True, exist_ok=True)
                docx_yolu = hedef_klasor / f"Kontrol_{sn}.docx"
                ilk = yukler[0]
                kontrol_dokumani_olustur(
                    sablon_path=_resource_path("Bosaltma_Kontrol_Sablonu.docx"),
                    cikti_path=docx_yolu,
                    tarih=ilk.tarih.strftime("%d.%m.%Y") if ilk.tarih else "",
                    gonderici_unvan=ilk.gonderen,
                    tasiyici_unvan=ilk.tasiyici_unvan,
                    plaka=plaka_sabit or ilk.plaka,
                    sefer_un_listesi=f"Sefer No: {sn} — {un_listesi}",
                    yangin_tup_tarihi=ek.get("yangin_tup", ""),
                    tmfb_tarihi=ek.get("tmfb", ""),
                    adr_uygunluk_tarihi=ek.get("adr_uygunluk", ""),
                    ara_muayene_tarihi=muayene_tarihleri.get(plaka_sabit, "") if muayene_tarihleri else "",
                    periyodik_muayene_tarihi=ek.get("periyodik_muayene", ""),
                    bosaltan_adi=bosaltan_adi,
                    sofor_adi=sofor_adi,
                    logo_bytes=logo_bytes,
                )
                docx_uretilen += 1
                emit(f"  📋 Kontrol dökümanı oluşturuldu → {docx_yolu.name}")

                pdf_yolu = None
                if docx_pdf_donustur:
                    pdf_yolu = docx_to_pdf(docx_yolu, hedef_klasor)
                    if pdf_yolu:
                        emit(f"  📑 PDF'e çevrildi → {pdf_yolu.name}")
                    else:
                        emit(f"  ⚠ PDF dönüşümü başarısız (LibreOffice bulunamadı/hata), .docx kullanılabilir.")

                # Excel'deki Sefer No hücresine eklenecek köprü: Excel çıktısının
                # bulunduğu klasöre göre GÖRECELİ yol (PDF varsa PDF'e, yoksa DOCX'e).
                hedef_dosya = pdf_yolu or docx_yolu
                try:
                    goreceli_yol = os.path.relpath(hedef_dosya, start=output_path.parent)
                except ValueError:
                    goreceli_yol = str(hedef_dosya)  # farklı sürücüde ise (Windows) mutlak yola düş
                for y in yukler:
                    y.kontrol_dokumani_yolu = goreceli_yol

                uretilen_dosyalar.append({"sefer_no": sn, "docx": docx_yolu, "pdf": pdf_yolu})
            except Exception as exc:
                emit(f"  ⚠ Kontrol dökümanı oluşturulamadı ({sn}): {exc}")
        elif docx_uret and not _DOCX_DESTEGI:
            emit("  ⚠ docx_doldur modülü/python-docx kurulu değil, kontrol dökümanı atlandı.")

    if not yeni:
        emit("Eklenecek yeni sefer bulunamadı.")
        return {"eklenen": 0, "atlanan": atlanan, "toplam": len(mevcut),
                "docx_uretilen": docx_uretilen, "uretilen_dosyalar": uretilen_dosyalar}

    tum = sorted(mevcut + yeni, key=lambda y: y.sort_key())
    _rewrite(ws, tum, style)

    if logo_bytes:
        try:
            _excel_logo_ekle(ws, logo_bytes)
        except Exception as exc:
            emit(f"  ⚠ Excel'e logo eklenemedi: {exc}")

    wb.save(output_path)
    emit(f"✅ Kaydedildi → {output_path.name}  ({eklenen} yeni sefer, {len(tum)} toplam satır)")
    return {"eklenen": eklenen, "atlanan": atlanan, "toplam": len(tum),
            "docx_uretilen": docx_uretilen, "uretilen_dosyalar": uretilen_dosyalar}


# ---------------------------------------------------------------------------
# Tkinter Arayüzü
# ---------------------------------------------------------------------------
