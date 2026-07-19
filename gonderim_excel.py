#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Atık gönderimlerini 1_Tasimacilik_Bilgi_Listesi formatında Excel'e yazar.

Sütun haritası (başlık satırı = 6, veri = 7+):
 1: Sıra No
 2: Taşıma Tarihi
 3: Taşımacı İşletme Unvanı
 4: Vergi Numarası           → boş (export'ta yok)
 5: Araç Plakası
 6: Taşıma Türü             → her zaman ADR-AMBALAJLI
 7: TMFB Geçerlilik Tarihi  → boş
 8: TMFB Numarası           → boş
 9: Tank/Ambalaj/Konteyner  → boş
10: Sertifika Geçerlilik    → boş
11: Yük Taşıma Birimi Muayene → boş
12: Basınçlı Kap Uygunluk   → İLGİLİ DEĞİL
13: Basınçlı Kap Muayene    → İLGİLİ DEĞİL
14: Taşıma Evrağı/İrsaliye  → Taşıma No(ları)
15: Muafiyet Kapsamında     → hesaplanır
16: SRC5 Belgeli Şoför      → boş
17: Yük Miktarı (KG)        → toplam kg
18: Kontrol Formunu Dolduran → kullanıcı adı (gonderici_adi)
"""
from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import io

from export_islem import AtikGonderim

# Sütun numaraları (1-based)
COL = {
    'sira_no': 1,
    'tarih': 2,
    'tasiyici': 3,
    'vergi_no': 4,
    'plaka': 5,
    'tasima_turu': 6,
    'tmfb_tarih': 7,
    'tmfb_no': 8,
    'sertifika_no': 9,
    'sertifika_tarih': 10,
    'muayene_tarih': 11,
    'basinc_uygun': 12,
    'basinc_muayene': 13,
    'evraki': 14,
    'muafiyet': 15,
    'src5': 16,
    'miktar': 17,
    'dolduran': 18,
}

DATA_START_ROW = 7
ZEBRA_RENKLER = ['D9E1F2', 'FFFFFF']  # Mavi/Beyaz zebra


def _zebra_fill(renk_hex: str) -> PatternFill:
    return PatternFill(fill_type='solid', fgColor=renk_hex)


def _template_style(ws) -> dict:
    """Veri başlangıç satırının üzerindeki stil referansını alır."""
    stil = {}
    ref_row = DATA_START_ROW
    # Şablonda ilk veri satırının stilini oku
    for c in range(1, 20):
        cell = ws.cell(row=ref_row, column=c)
        stil[c] = {
            'font': copy(cell.font),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format,
        }
    return stil


def _read_existing(ws) -> tuple[list[dict], set[str]]:
    """Mevcut Excel'deki veri satırlarını ve taşıma no listesini döndürür."""
    satirlar = []
    tasima_nolar = set()
    for row in range(DATA_START_ROW, ws.max_row + 1):
        evraki = ws.cell(row=row, column=COL['evraki']).value
        if not evraki:
            break
        evraki_str = str(evraki).strip()
        satirlar.append(row)
        # Taşıma No'ları ayır (virgülle birleştirilmiş olabilir)
        for no in evraki_str.split(','):
            tasima_nolar.add(no.strip())
    return satirlar, tasima_nolar


def _rewrite(ws, gonderimler: list[AtikGonderim], stil: dict,
             gonderici_adi: str = '', logo_bytes: Optional[bytes] = None) -> None:
    """Tüm veri satırlarını siler ve gonderimler listesinden yeniden yazar."""
    # Mevcut veri satırlarını temizle
    for row in range(ws.max_row, DATA_START_ROW - 1, -1):
        if ws.cell(row=row, column=1).value is not None:
            ws.delete_rows(row)

    # Logo
    if logo_bytes:
        if hasattr(ws, '_images'):
            ws._images = []
        img = XLImage(io.BytesIO(logo_bytes))
        img.width = 140
        img.height = 70
        ws.add_image(img, 'A1')

    zebra_map = {}
    zebra_idx = 0
    for g in gonderimler:
        grup_key = g.tasima_nolari_str
        if grup_key not in zebra_map:
            zebra_map[grup_key] = ZEBRA_RENKLER[zebra_idx % len(ZEBRA_RENKLER)]
            zebra_idx += 1

    for sira, g in enumerate(gonderimler, start=1):
        row = DATA_START_ROW + sira - 1
        fill = _zebra_fill(zebra_map[g.tasima_nolari_str])

        degerler = {
            COL['sira_no']: sira,
            COL['tarih']: g.tarih,
            COL['tasiyici']: g.tasiyici,
            COL['vergi_no']: None,
            COL['plaka']: g.plaka,
            COL['tasima_turu']: 'ADR-AMBALAJLI',
            COL['tmfb_tarih']: None,
            COL['tmfb_no']: None,
            COL['sertifika_no']: g.atik_kodlari_str,
            COL['sertifika_tarih']: None,
            COL['muayene_tarih']: None,
            COL['basinc_uygun']: 'İLGİLİ DEĞİL',
            COL['basinc_muayene']: 'İLGİLİ DEĞİL',
            COL['evraki']: g.tasima_nolari_str,
            COL['muafiyet']: g.muafiyet,
            COL['src5']: None,
            COL['miktar']: g.miktar_kg,
            COL['dolduran']: gonderici_adi,
        }

        for col, deger in degerler.items():
            cell = ws.cell(row=row, column=col, value=deger)
            if col in stil:
                s = stil[col]
                if s['font']:
                    cell.font = copy(s['font'])
                if s['border']:
                    cell.border = copy(s['border'])
                if s['alignment']:
                    cell.alignment = copy(s['alignment'])
                if s['number_format']:
                    cell.number_format = s['number_format']
            cell.fill = fill

        ws.row_dimensions[row].height = 45


def process_export(
    sablon_excel_path: str | Path,
    gonderimler: list[AtikGonderim],
    cikti_path: str | Path,
    gonderici_adi: str = '',
    logo_bytes: Optional[bytes] = None,
    log_cb=None,
) -> dict:
    """
    Gönderim verilerini Excel şablonuna yazar.

    Returns:
        {'eklenen': int, 'atlanan': int, 'toplam': int}
    """
    def emit(msg):
        if log_cb:
            log_cb(msg)

    wb = openpyxl.load_workbook(str(sablon_excel_path))
    ws = wb.active
    stil = _template_style(ws)
    mevcut_satirlar, mevcut_tasima_nolar = _read_existing(ws)

    yeni = []
    atlanan = 0
    for g in gonderimler:
        # Herhangi bir taşıma no'su zaten varsa atla
        cakisan = any(no in mevcut_tasima_nolar for no in g.tasima_nolari)
        if cakisan:
            emit(f'  ⚠ {g.tasima_nolari_str} zaten mevcut, atlandı.')
            atlanan += 1
        else:
            yeni.append(g)
            emit(f'  ✓ {g.tasima_nolari_str} — {g.tarih_str} / {g.plaka} ({g.miktar_kg:.0f} kg)')

    # Mevcut satırları da oku (yeniden sıralama için)
    mevcut_gonderimler = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        if not ws.cell(row=row, column=COL['evraki']).value:
            break
        tarih_val = ws.cell(row=row, column=COL['tarih']).value
        tasiyici_val = ws.cell(row=row, column=COL['tasiyici']).value or ''
        plaka_val = ws.cell(row=row, column=COL['plaka']).value or ''
        evraki_val = str(ws.cell(row=row, column=COL['evraki']).value or '')
        miktar_val = ws.cell(row=row, column=COL['miktar']).value or 0
        muafiyet_val = ws.cell(row=row, column=COL['muafiyet']).value or ''

        g_mevcut = AtikGonderim(
            tarih=tarih_val if isinstance(tarih_val, datetime) else None,
            tasiyici=str(tasiyici_val),
            plaka=str(plaka_val),
            alici='',
            atik_kodlari=[str(ws.cell(row=row, column=COL['sertifika_no']).value or '')],
            tasima_nolari=evraki_val.split(', '),
            un_nolar=[],
            miktar_kg=float(miktar_val),
        )
        mevcut_gonderimler.append(g_mevcut)

    tum = sorted(mevcut_gonderimler + yeni, key=lambda g: g.tarih or datetime.min)
    _rewrite(ws, tum, stil, gonderici_adi=gonderici_adi, logo_bytes=logo_bytes)

    cikti_path = Path(cikti_path)
    cikti_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(cikti_path))
    emit(f'✅ Excel kaydedildi → {cikti_path.name} ({len(yeni)} yeni, {len(tum)} toplam)')
    return {'eklenen': len(yeni), 'atlanan': atlanan, 'toplam': len(tum)}
